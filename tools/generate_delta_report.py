#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SISL: Price Delta Report "SISL_VFD_PRICE_DELTA_2024_vs_2025_v01"
- Columns: Capacity (kW), Year, Model, Price (BDT), Î” (BDT), Î” (%), Notes
- 2 rows per model group (2024 first, 2025 second), same pairing logic as v01
- Keep 2025 model names exactly
- Alternating group shading (white / light grey); dark grey if any price missing
"""
import os, re, sys, argparse
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

OUTDEFAULT = r"reports\SISL_VFD_PRICE_DELTA_2024_vs_2025_v01"
CAP_CAND   = ["Capacity (KW)", "Capacity (kW)", "Capacity", "kW"]
MODEL_CAND = ["Model (Rated Current)", "Model", "MODEL", "Model Name"]
PRICE_CAND = ["Price (BDT)", "LP (BDT)", "List Price", "LP_BDT", "Price", "COGS"]

def pick_col(df, options):
    for c in options:
        if c in df.columns: return c
    return None

def cap_to_float(s):
    s = str(s).strip()
    if s == "" or s.lower() in ("nan", "none"): return float("nan")
    s = re.sub(r"\s*(kW|KW)\b", "", s)
    if s.endswith("K"): s = s[:-1]
    try: return float(s)
    except ValueError:
        m = re.search(r"(\d+(?:\.\d+)?)", s)
        return float(m.group(1)) if m else float("nan")

def price_to_float(x):
    s = re.sub(r"[^\d.\-]", "", str(x))
    try: return float(s) if s not in ("", ".", "-", "--") else float("nan")
    except ValueError: return float("nan")

def read_csv(path):
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    cap = pick_col(df, CAP_CAND); mod = pick_col(df, MODEL_CAND); prc = pick_col(df, PRICE_CAND)
    if not (cap and mod and prc):
        raise RuntimeError(f"{path}: missing required columns. Found: {list(df.columns)}")
    out = pd.DataFrame({
        "Capacity (kW)": df[cap].map(cap_to_float),
        "Model": df[mod],
        "Price (BDT)": df[prc],
        "_PriceNum": df[prc].map(price_to_float),
    })
    def series_of(m):
        m = str(m)
        m2 = re.search(r"FR-([A-Z]+)", m)
        if not m2: return "Z"
        s = m2.group(1)
        if s.startswith("HEL"): return "HEL"
        if s.startswith("CS"):  return "D"
        return s[0]
    out["Series"] = out["Model"].map(series_of)
    return out

def autodetect(year):
    for folder in [Path("data"), Path("."), Path("Datasets"), Path("dataset")]:
        if folder.exists():
            for p in sorted(folder.glob("*.csv")):
                if str(year) in p.name: return str(p)
    return None

def parse_args():
    p = argparse.ArgumentParser(description="Generate price-delta report 2024 vs 2025")
    p.add_argument("--y2024", default=None, help="Path to 2024 CSV")
    p.add_argument("--y2025", default=None, help="Path to 2025 CSV")
    p.add_argument("--outprefix", default=None, help="Override output prefix path")
    return p.parse_args()

def main():
    args = parse_args()
    y2024 = args.y2024 or os.environ.get("Y2024_CSV") or autodetect(2024)
    y2025 = args.y2025 or os.environ.get("Y2025_CSV") or autodetect(2025)
    outprefix = args.outprefix or os.environ.get("OUTPREFIX") or OUTDEFAULT

    if not y2024 or not Path(y2024).exists():
        print("ERROR: 2024 CSV not found. Put a *2024*.csv in .\\data or set --y2024", file=sys.stderr); sys.exit(1)
    if not y2025 or not Path(y2025).exists():
        print("ERROR: 2025 CSV not found. Put a *2025*.csv in .\\data or set --y2025", file=sys.stderr); sys.exit(1)

    df24 = read_csv(y2024).copy()
    df25 = read_csv(y2025).copy()

    # anchor by 2024 appearance
    df24["_idx"] = range(len(df24))
    df24_sorted = df24.sort_values(["Capacity (kW)", "_idx"])

    # match by (capacity, series) first-unused
    df25["_used"] = False
    pairs = []
    for _, r24 in df24_sorted.iterrows():
        cap = r24["Capacity (kW)"]; ser = r24["Series"]
        cand = df25[(df25["Capacity (kW)"] == cap) & (df25["Series"] == ser) & (~df25["_used"])]
        r25 = cand.iloc[0] if not cand.empty else None
        if r25 is not None: df25.loc[r25.name, "_used"] = True
        pairs.append((r24, r25))

    # 2025-only leftovers
    unused25 = df25[~df25["_used"]].copy().sort_values(["Capacity (kW)", "Series", "Model"])

    cap_24 = list(df24_sorted["Capacity (kW)"].dropna().unique())
    cap_only25 = sorted(set(unused25["Capacity (kW)"].dropna().unique()) - set(cap_24))
    cap_all = sorted(cap_24 + cap_only25)

    groups = []
    for cap in cap_all:
        groups.extend([(r24, r25) for (r24, r25) in pairs if r24["Capacity (kW)"] == cap])
        for _, r25 in unused25[unused25["Capacity (kW)"] == cap].iterrows():
            groups.append((None, r25))
    # unknown-capacity 2025-only at end
    for _, r25 in unused25[unused25["Capacity (kW)"].isna()].iterrows():
        groups.append((None, r25))

    # build rows (two per group)
    rows = []
    for r24, r25 in groups:
        cap_val = r24["Capacity (kW)"] if r24 is not None else (r25["Capacity (kW)"] if r25 is not None else float("nan"))
        p24 = (r24["_PriceNum"] if r24 is not None else float("nan"))
        p25 = (r25["_PriceNum"] if r25 is not None else float("nan"))
        dlt = (p25 - p24) if (pd.notna(p25) and pd.notna(p24)) else float("nan")
        pct = (dlt / p24) if (pd.notna(dlt) and pd.notna(p24) and p24 != 0) else float("nan")
        rows.append({
            "Capacity (kW)": cap_val, "Year": 2024,
            "Model": (r24["Model"] if r24 is not None else ""),
            "Price (BDT)": (r24["Price (BDT)"] if r24 is not None else ""),
            "Î” (BDT)": "", "Î” (%)": "", "Notes": ("" if r25 is not None else "Missing 2025"),
            "_p24": p24, "_p25": p25,
        })
        rows.append({
            "Capacity (kW)": cap_val, "Year": 2025,
            "Model": (r25["Model"] if r25 is not None else ""),
            "Price (BDT)": (r25["Price (BDT)"] if r25 is not None else ""),
            "Î” (BDT)": dlt, "Î” (%)": pct, "Notes": ("" if r24 is not None else "Missing 2024"),
            "_p24": p24, "_p25": p25,
        })

    out_cols = ["Capacity (kW)", "Year", "Model", "Price (BDT)", "Î” (BDT)", "Î” (%)", "Notes"]
    df = pd.DataFrame(rows, columns=out_cols + ["_p24","_p25"])

    # CSV
    Path(outprefix).parent.mkdir(parents=True, exist_ok=True)
    csv_path = Path(f"{outprefix}.csv"); df[out_cols].to_csv(csv_path, index=False, encoding="utf-8-sig")

    # Excel
    wb = Workbook(); ws = wb.active; ws.title = "Price Delta Report"
    headers = out_cols; ws.append(headers)
    light_grey = PatternFill("solid", fgColor="B3B3B3")
    dark_grey  = PatternFill("solid", fgColor="999999")
    hdr_fill   = PatternFill("solid", fgColor="DDDDDD")
    thin = Side(style="thin", color="BBBBBB"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True); center = Alignment(horizontal="center", vertical="center")
    left  = Alignment(horizontal="left",  vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold; cell.alignment = center; cell.fill = hdr_fill; cell.border = border

    r = 2
    for i in range(0, len(df), 2):
        pair = df.iloc[i:i+2]
        any_missing_price = (pd.isna(pair["_p24"]).any() or pd.isna(pair["_p25"]).any())
        row_fill = dark_grey if any_missing_price else (light_grey if ((i // 2) % 2 == 1) else None)
        for _, row in pair.iterrows():
            ws.append([row[h] for h in headers])
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=c); cell.border = border
                if h in ("Year", "Capacity (kW)"): cell.alignment = center
                elif h in ("Model", "Notes"): cell.alignment = left
                elif h in ("Î” (BDT)", "Price (BDT)"): cell.alignment = right
                else: cell.alignment = center
                if row_fill: cell.fill = row_fill
                if h == "Î” (BDT)" and isinstance(row["Î” (BDT)"], (int,float)) and pd.notna(row["Î” (BDT)"]): cell.number_format = '#,##0'
                if h == "Price (BDT)" and isinstance(row["Price (BDT)"], (int,float)): cell.number_format = '#,##0'
                if h == "Î” (%)" and isinstance(row["Î” (%)"], (int,float)) and pd.notna(row["Î” (%)"]): cell.number_format = '0.0%'
            r += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 30
    ws.freeze_panes = "A2"

    last_row = ws.max_row
    rng = f"E2:E{last_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=False,
        fill=PatternFill(start_color='F8CECC', end_color='F8CECC', fill_type='solid')))
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=False,
        fill=PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')))

    xlsx_path = Path(f"{outprefix}.xlsx"); wb.save(xlsx_path)
    print(f"âœ“ Wrote: {csv_path}"); print(f"âœ“ Wrote: {xlsx_path}")

if __name__ == "__main__":
    main()

