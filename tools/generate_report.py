#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate "SISL_VFD_PA_2024 vs 2025_v01" (CSV + Excel) from repo CSVs.

- Output columns: Capacity (kW), Year, Model, Price (BDT)
- Capacity ascending (0.4 → … → 400)
- Exactly 2 rows per group: 2024 first, 2025 second
- Within each capacity: follow 2024 order; then append any 2025-only models
- Keep 2025 Model names EXACTLY as in the 2025 CSV
- Excel: strict two-tone group banding (White/Grey #B3B3B3), header bold, widths A=14 B=8 C=60 D=18
"""
import os, re, sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DEFAULT_OUTPREFIX = r"reports\SISL_VFD_PA_2024 vs 2025_v01"
CAP_CAND   = ["Capacity (KW)", "Capacity (kW)", "Capacity", "kW"]
MODEL_CAND = ["Model (Rated Current)", "Model", "MODEL", "Model Name"]
PRICE_CAND = ["Price (BDT)", "LP (BDT)", "List Price", "LP_BDT", "Price", "COGS"]

def pick_col(df, options):
    for c in options:
        if c in df.columns:
            return c
    return None

def cap_to_float(s):
    """Tolerant capacity parser: '', NaN → NaN; strips 'K' and 'kW' text."""
    s = str(s).strip()
    if s == "" or s.lower() in ("nan", "none"):
        return float("nan")
    s = re.sub(r"\s*(kW|KW)\b", "", s)  # drop 'kW' suffix if present
    if s.endswith("K"):
        s = s[:-1]
    try:
        return float(s)
    except ValueError:
        m = re.search(r"(\d+(?:\.\d+)?)", s)
        return float(m.group(1)) if m else float("nan")

def read_csv(path):
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    cap = pick_col(df, CAP_CAND)
    mod = pick_col(df, MODEL_CAND)
    prc = pick_col(df, PRICE_CAND)
    if not (cap and mod and prc):
        raise RuntimeError(f"{path}: missing required columns. Found: {list(df.columns)}")
    out = pd.DataFrame({
        "Capacity (kW)": df[cap].map(cap_to_float),
        "Model": df[mod],        # keep verbatim
        "Price (BDT)": df[prc],  # keep as text
    })
    # Series group for pairing (D/E/F/A). CS→D, HEL stays HEL. Fallback 'Z'.
    def series_of(m):
        m = str(m)
        m2 = re.search(r"FR-([A-Z]+)", m)
        if not m2: return "Z"
        s = m2.group(1)
        if s.startswith("HEL"): return "HEL"
        if s.startswith("CS"):  return "D"
        return s[0]
    out["Series"] = out["Model"].map(series_of)
    return out

def autodetect(year):
    for folder in [Path("data"), Path("."), Path("Datasets"), Path("dataset")]:
        if folder.exists():
            for p in sorted(folder.glob("*.csv")):
                if str(year) in p.name:
                    return str(p)
    return None

def main():
    # Inputs (auto-detect or environment overrides)
    y2024 = os.environ.get("Y2024_CSV") or autodetect(2024)
    y2025 = os.environ.get("Y2025_CSV") or autodetect(2025)
    outprefix = os.environ.get("OUTPREFIX") or DEFAULT_OUTPREFIX

    if not y2024 or not Path(y2024).exists():
        print("ERROR: 2024 CSV not found. Put a *2024*.csv in .\\data or set Y2024_CSV", file=sys.stderr); sys.exit(1)
    if not y2025 or not Path(y2025).exists():
        print("ERROR: 2025 CSV not found. Put a *2025*.csv in .\\data or set Y2025_CSV", file=sys.stderr); sys.exit(1)

    df24 = read_csv(y2024).copy()
    df25 = read_csv(y2025).copy()

    # Anchor by 2024 appearance per capacity
    df24["_idx"] = range(len(df24))
    df24_sorted = df24.sort_values(["Capacity (kW)", "_idx"])

    # Pair 2024→2025 by same capacity & series (first unused)
    df25["_used"] = False
    pairs = []
    for _, r24 in df24_sorted.iterrows():
        cap = r24["Capacity (kW)"]
        ser = r24["Series"]
        cand = df25[
            (df25["Capacity (kW)"] == cap) &
            (df25["Series"] == ser) &
            (~df25["_used"])
        ]
        r25 = cand.iloc[0] if not cand.empty else None
        if r25 is not None:
            df25.loc[r25.name, "_used"] = True
        pairs.append((r24, r25))

    # 2025-only rows (not matched to any 2024)
    unused25 = df25[~df25["_used"]].copy().sort_values(["Capacity (kW)", "Series", "Model"])

    # Capacity order: all 2024 capacities, then 2025-only capacities (NaN excluded here)
    cap_order_24 = df24_sorted["Capacity (kW)"].dropna().unique().tolist()
    cap_only_25 = sorted(set(unused25["Capacity (kW)"].dropna().unique()) - set(cap_order_24))
    cap_order_all = sorted(cap_order_24 + cap_only_25)

    # Build groups in desired order
    group_pairs = []
    for cap in cap_order_all:
        # 24-anchored groups for this capacity
        group_pairs.extend([(r24, r25) for (r24, r25) in pairs if r24["Capacity (kW)"] == cap])
        # 2025-only at this capacity
        for _, r25 in unused25[unused25["Capacity (kW)"] == cap].iterrows():
            group_pairs.append((None, r25))
    # Finally, append any 2025-only rows with unknown capacity
    for _, r25 in unused25[unused25["Capacity (kW)"].isna()].iterrows():
        group_pairs.append((None, r25))

    # Emit rows (two per group: 2024 then 2025)
    rows = []
    for r24, r25 in group_pairs:
        cap_24 = r24["Capacity (kW)"] if r24 is not None else (r25["Capacity (kW)"] if r25 is not None else float("nan"))
        cap_25 = r25["Capacity (kW)"] if r25 is not None else (r24["Capacity (kW)"] if r24 is not None else float("nan"))
        rows.append({
            "Capacity (kW)": cap_24,
            "Year": 2024,
            "Model": (r24["Model"] if r24 is not None else ""),
            "Price (BDT)": (r24["Price (BDT)"] if r24 is not None else ""),
        })
        rows.append({
            "Capacity (kW)": cap_25,
            "Year": 2025,
            "Model": (r25["Model"] if r25 is not None else ""),
            "Price (BDT)": (r25["Price (BDT)"] if r25 is not None else ""),
        })

    out_df = pd.DataFrame(rows, columns=["Capacity (kW)", "Year", "Model", "Price (BDT)"])

    # Save CSV
    Path(outprefix).parent.mkdir(parents=True, exist_ok=True)
    csv_path = Path(f"{outprefix}.csv")
    out_df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    # Save Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "First Report"
    headers = ["Capacity (kW)", "Year", "Model", "Price (BDT)"]
    ws.append(headers)

    grey = PatternFill("solid", fgColor="B3B3B3")
    hdr = PatternFill("solid", fgColor="DDDDDD")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Header styling
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.alignment = center
        cell.fill = hdr
        cell.border = border

    # Strict alternation: Group1 white, Group2 grey, Group3 white, ...
    r = 2
    for i in range(0, len(out_df), 2):
        pair = out_df.iloc[i:i+2]
        fill = grey if ((i // 2) % 2 == 1) else None
        for _, row in pair.iterrows():
            ws.append([row[h] for h in headers])
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if h == "Year":
                    cell.alignment = center
                elif h == "Model":
                    cell.alignment = left
                elif h == "Price (BDT)":
                    cell.alignment = right
                else:
                    cell.alignment = center
                if fill:
                    cell.fill = fill
            r += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A2"

    xlsx_path = Path(f"{outprefix}.xlsx")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)

    print(f"✓ Wrote: {csv_path}")
    print(f"✓ Wrote: {xlsx_path}")

if __name__ == "__main__":
    main()
